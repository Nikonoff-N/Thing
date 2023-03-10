import pandas as pd
from scraper import ProductData
import openpyxl
from datetime import datetime
FILENAME = 'Data.xlsx'
GREEN = '000000FF'
RED = '00FF0000'
BLUE = '0000FF00'
def write_data(data:list[ProductData])->None:
    date = datetime.today().strftime('%Y-%m-%d')

    # Load the existing Excel file
    wb = openpyxl.load_workbook(FILENAME)
    # Get the existing sheet in the Excel file
    ws = wb['Data']
    # Add 2 new columns to the sheet
    # ws.insert_cols(ws.max_column + 1, 2)
    ws.cell(column=ws.max_column+1,row = 1).value = date
    ws.cell(column=ws.max_column,row = 2).value = "Цена"
    ws.cell(column=ws.max_column+1,row = 2).value = "Кол-во"
    for row,info in zip(ws.iter_rows(min_row=3,min_col=ws.max_column-1, max_col=ws.max_column, max_row=ws.max_row),data):
        if info.state == "A":
            row[0].value = info.price
            row[1].value = info.count
            ws.cell(column = 3,row = row[0].row).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor= openpyxl.styles.colors.Color(rgb=GREEN))
            ws.cell(column = 1,row = row[0].row).value = info.name
        elif info.state == "B":
            ws.cell(column = 3,row = row[0].row).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor= openpyxl.styles.colors.Color(rgb=RED))
        else:
            ws.cell(column = 3,row = row[0].row).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor= openpyxl.styles.colors.Color(rgb=BLUE))
        ws.cell(column = 3,row = row[0].row).value = info.state
    wb.save(FILENAME)